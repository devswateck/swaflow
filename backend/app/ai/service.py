import csv
import json
from io import BytesIO
from copy import deepcopy
from uuid import UUID

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.ai.intent_classifier import classify_intent
from app.ai.models import AiAgent, AiFaqEntry, AiInteractiveTemplate
from app.ai.schemas import (
    AiAgentCreate,
    AiAgentUpdate,
    AiFaqEntryCreate,
    AiFaqEntryUpdate,
    AiFaqUploadResult,
    AiInteractiveTemplateCreate,
    AiInteractiveTemplateUpdate,
    AiOperationalConfigRead,
    AiOperationalSimulationResponse,
    IntentClassifyResponse,
)
from app.ai.operational import (
    build_operational_config,
    publish_operational_config,
    simulation_summary,
    validate_operational_config,
)
from app.audit.service import record_audit_best_effort
from app.companies.models import Company
from app.users.models import User

MAX_FAQ_ENTRIES_PER_TENANT = 10


def _audit_ai_action(
    db: Session,
    *,
    company_id: UUID,
    actor_user: User | None,
    action: str,
    entity_type: str,
    entity_id: UUID | None,
    summary: str,
    metadata: dict | None = None,
) -> None:
    record_audit_best_effort(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        summary=summary,
        metadata=metadata or {},
    )


def _agent_rules_copy(agent: AiAgent | None) -> dict:
    if agent is None or not isinstance(agent.rules, dict):
        return {}
    return deepcopy(agent.rules)


def _normalize_agent_rules(
    *,
    base_rules: dict | None,
    operational_config: dict | None,
    fallback_timezone: str | None,
) -> dict:
    rules = deepcopy(base_rules) if isinstance(base_rules, dict) else {}
    effective_operational = build_operational_config(rules, fallback_timezone=fallback_timezone)
    if operational_config is not None:
        if not isinstance(operational_config, dict):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="La configuracion operativa debe ser un objeto valido.",
            )
        effective_operational = build_operational_config(
            {"operational": operational_config},
            fallback_timezone=fallback_timezone,
        )
    validate_operational_config(effective_operational)
    rules["operational"] = effective_operational
    return rules


def _agent_timezone(db: Session, *, company_id: UUID) -> str | None:
    company = db.scalar(select(Company).where(Company.id == company_id))
    if company is None:
        return None
    return company.timezone


def _agent_operational_config(db: Session, *, company_id: UUID, agent: AiAgent) -> dict:
    fallback_timezone = _agent_timezone(db, company_id=company_id)
    return build_operational_config(agent.rules if isinstance(agent.rules, dict) else {}, fallback_timezone=fallback_timezone)


def list_agents(db: Session, *, company_id: UUID) -> list[AiAgent]:
    # Single-agent mode per tenant: return only the canonical record.
    agent = db.scalar(
        select(AiAgent)
        .where(AiAgent.company_id == company_id)
        .order_by(AiAgent.updated_at.desc(), AiAgent.created_at.desc())
    )
    return [agent] if agent else []


def _cleanup_extra_agents(db: Session, *, company_id: UUID, keep_id: UUID) -> None:
    extras = list(
        db.scalars(
            select(AiAgent).where(
                AiAgent.company_id == company_id,
                AiAgent.id != keep_id,
            )
        )
    )
    for extra in extras:
        db.delete(extra)


def create_agent(
    db: Session,
    *,
    company_id: UUID,
    payload: AiAgentCreate,
    actor_user: User | None = None,
) -> AiAgent:
    existing = db.scalar(
        select(AiAgent)
        .where(AiAgent.company_id == company_id)
        .order_by(AiAgent.updated_at.desc(), AiAgent.created_at.desc())
        .with_for_update()
    )
    fallback_timezone = _agent_timezone(db, company_id=company_id)
    normalized_rules = _normalize_agent_rules(
        base_rules=payload.rules,
        operational_config=payload.operational_config,
        fallback_timezone=fallback_timezone,
    )
    try:
        if existing is None:
            agent = AiAgent(
                company_id=company_id,
                name=payload.name,
                system_prompt=payload.system_prompt,
                conversation_objective=payload.conversation_objective,
                conversation_guide=payload.conversation_guide,
                security_rules=payload.security_rules,
                tone=payload.tone,
                rules=normalized_rules,
                active=payload.active,
            )
            db.add(agent)
            db.flush()
        else:
            existing.name = payload.name
            existing.system_prompt = payload.system_prompt
            existing.conversation_objective = payload.conversation_objective
            existing.conversation_guide = payload.conversation_guide
            existing.security_rules = payload.security_rules
            existing.tone = payload.tone
            existing.rules = normalized_rules
            existing.active = payload.active
            agent = existing

        _cleanup_extra_agents(db, company_id=company_id, keep_id=agent.id)
        db.commit()
        db.refresh(agent)
        _audit_ai_action(
            db,
            company_id=company_id,
            actor_user=actor_user,
            action="ai.agent.saved",
            entity_type="ai_agent",
            entity_id=agent.id,
            summary="AI agent saved",
            metadata={"name": agent.name, "active": agent.active},
        )
        return agent
    except IntegrityError:
        db.rollback()
        agent = db.scalar(
            select(AiAgent)
            .where(AiAgent.company_id == company_id)
            .order_by(AiAgent.updated_at.desc(), AiAgent.created_at.desc())
        )
        if agent is None:
            raise
        agent.name = payload.name
        agent.system_prompt = payload.system_prompt
        agent.conversation_objective = payload.conversation_objective
        agent.conversation_guide = payload.conversation_guide
        agent.security_rules = payload.security_rules
        agent.tone = payload.tone
        agent.rules = normalized_rules
        agent.active = payload.active
        _cleanup_extra_agents(db, company_id=company_id, keep_id=agent.id)
        db.commit()
        db.refresh(agent)
        _audit_ai_action(
            db,
            company_id=company_id,
            actor_user=actor_user,
            action="ai.agent.saved",
            entity_type="ai_agent",
            entity_id=agent.id,
            summary="AI agent saved",
            metadata={"name": agent.name, "active": agent.active},
        )
        return agent


def get_agent(db: Session, *, company_id: UUID, agent_id: UUID) -> AiAgent:
    agent = db.scalar(select(AiAgent).where(AiAgent.company_id == company_id, AiAgent.id == agent_id))
    if agent is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="AI agent not found")
    return agent


def update_agent(
    db: Session,
    *,
    company_id: UUID,
    agent_id: UUID,
    payload: AiAgentUpdate,
    actor_user: User | None = None,
) -> AiAgent:
    agent = get_agent(db, company_id=company_id, agent_id=agent_id)
    data = payload.model_dump(exclude_unset=True)
    fallback_timezone = _agent_timezone(db, company_id=company_id)
    if "rules" in data or "operational_config" in data:
        base_rules = data.get("rules") if data.get("rules") is not None else _agent_rules_copy(agent)
        data["rules"] = _normalize_agent_rules(
            base_rules=base_rules,
            operational_config=data.get("operational_config"),
            fallback_timezone=fallback_timezone,
        )
        data.pop("operational_config", None)
    for field, value in data.items():
        setattr(agent, field, value)
    _cleanup_extra_agents(db, company_id=company_id, keep_id=agent.id)
    db.commit()
    db.refresh(agent)
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.agent.updated",
        entity_type="ai_agent",
        entity_id=agent.id,
        summary="AI agent updated",
        metadata={"name": agent.name, "active": agent.active},
    )
    return agent


def classify_message(message: str) -> IntentClassifyResponse:
    return classify_intent(message)


def list_faq_entries(db: Session, *, company_id: UUID) -> list[AiFaqEntry]:
    return list(
        db.scalars(
            select(AiFaqEntry)
            .where(AiFaqEntry.company_id == company_id)
            .order_by(AiFaqEntry.created_at.asc())
        )
    )


def _count_faq_entries(db: Session, *, company_id: UUID) -> int:
    return len(list_faq_entries(db, company_id=company_id))


def _normalize_faq_payload(question: str, answer: str) -> tuple[str, str]:
    return question.strip(), answer.strip()


def create_faq_entry(
    db: Session,
    *,
    company_id: UUID,
    payload: AiFaqEntryCreate,
    actor_user: User | None = None,
) -> AiFaqEntry:
    if _count_faq_entries(db, company_id=company_id) >= MAX_FAQ_ENTRIES_PER_TENANT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo puedes guardar hasta {MAX_FAQ_ENTRIES_PER_TENANT} preguntas frecuentes.",
        )
    question, answer = _normalize_faq_payload(payload.question, payload.answer)
    entry = AiFaqEntry(
        company_id=company_id,
        question=question,
        answer=answer,
        active=payload.active,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.faq_entry.created",
        entity_type="ai_faq_entry",
        entity_id=entry.id,
        summary="FAQ entry created",
        metadata={"question": entry.question, "active": entry.active},
    )
    return entry


def get_faq_entry(db: Session, *, company_id: UUID, faq_id: UUID) -> AiFaqEntry:
    entry = db.scalar(
        select(AiFaqEntry).where(
            AiFaqEntry.company_id == company_id,
            AiFaqEntry.id == faq_id,
        )
    )
    if entry is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="FAQ not found")
    return entry


def update_faq_entry(
    db: Session,
    *,
    company_id: UUID,
    faq_id: UUID,
    payload: AiFaqEntryUpdate,
    actor_user: User | None = None,
) -> AiFaqEntry:
    entry = get_faq_entry(db, company_id=company_id, faq_id=faq_id)
    data = payload.model_dump(exclude_unset=True)
    if "question" in data or "answer" in data:
        question = data.get("question", entry.question)
        answer = data.get("answer", entry.answer)
        entry.question, entry.answer = _normalize_faq_payload(question, answer)
        data.pop("question", None)
        data.pop("answer", None)
    for field, value in data.items():
        setattr(entry, field, value)
    db.commit()
    db.refresh(entry)
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.faq_entry.updated",
        entity_type="ai_faq_entry",
        entity_id=entry.id,
        summary="FAQ entry updated",
        metadata={"question": entry.question, "active": entry.active},
    )
    return entry


def delete_faq_entry(
    db: Session,
    *,
    company_id: UUID,
    faq_id: UUID,
    actor_user: User | None = None,
) -> None:
    entry = get_faq_entry(db, company_id=company_id, faq_id=faq_id)
    entry_id = entry.id
    db.delete(entry)
    db.commit()
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.faq_entry.deleted",
        entity_type="ai_faq_entry",
        entity_id=entry_id,
        summary="FAQ entry deleted",
        metadata={"question": entry.question},
    )


def _decode_upload(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="No fue posible leer el archivo. Usa UTF-8, CSV, JSON o XLSX.",
    )


def _parse_csv_rows(content: bytes) -> list[tuple[str, str]]:
    text = _decode_upload(content)
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return []
    sample = "\n".join(lines[:10])
    delimiter = ";"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
        delimiter = dialect.delimiter
    except csv.Error:
        if "," in sample and ";" not in sample:
            delimiter = ","
    rows: list[tuple[str, str]] = []
    reader = csv.reader(lines, delimiter=delimiter)
    parsed = list(reader)
    if not parsed:
        return rows
    header = [cell.strip().lower() for cell in parsed[0]]
    has_header = bool({"question", "pregunta"} & set(header)) and bool(
        {"answer", "respuesta"} & set(header)
    )
    if has_header:
        question_index = next(
            i for i, value in enumerate(header) if value in {"question", "pregunta"}
        )
        answer_index = next(
            i for i, value in enumerate(header) if value in {"answer", "respuesta"}
        )
        data_rows = parsed[1:]
    else:
        question_index = 0
        answer_index = 1
        data_rows = parsed
    for row in data_rows:
        if len(row) <= max(question_index, answer_index):
            continue
        question = row[question_index].strip()
        answer = row[answer_index].strip()
        if question and answer:
            rows.append((question, answer))
    return rows


def _parse_json_rows(content: bytes) -> list[tuple[str, str]]:
    text = _decode_upload(content)
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"JSON invalido: {exc.msg}",
        ) from exc
    if not isinstance(payload, list):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El JSON debe ser una lista de objetos con question/answer.",
        )
    rows: list[tuple[str, str]] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        question = str(item.get("question") or item.get("pregunta") or "").strip()
        answer = str(item.get("answer") or item.get("respuesta") or "").strip()
        if question and answer:
            rows.append((question, answer))
    return rows


def _parse_xlsx_rows(content: bytes) -> list[tuple[str, str]]:
    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No fue posible leer el XLSX: {exc}",
        ) from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(cell or "").strip().lower() for cell in rows[0]]
    question_index = 0
    answer_index = 1
    has_header = bool({"question", "pregunta"} & set(header)) and bool(
        {"answer", "respuesta"} & set(header)
    )
    data_rows = rows[1:] if has_header else rows
    if has_header:
        question_index = next(
            i for i, value in enumerate(header) if value in {"question", "pregunta"}
        )
        answer_index = next(
            i for i, value in enumerate(header) if value in {"answer", "respuesta"}
        )
    parsed: list[tuple[str, str]] = []
    for row in data_rows:
        if row is None:
            continue
        if len(row) <= max(question_index, answer_index):
            continue
        question = str(row[question_index] or "").strip()
        answer = str(row[answer_index] or "").strip()
        if question and answer:
            parsed.append((question, answer))
    return parsed


def _parse_faq_file(filename: str, content: bytes) -> list[tuple[str, str]]:
    normalized = filename.lower()
    if normalized.endswith(".json"):
        return _parse_json_rows(content)
    if normalized.endswith(".xlsx"):
        return _parse_xlsx_rows(content)
    if normalized.endswith(".csv") or normalized.endswith(".txt"):
        return _parse_csv_rows(content)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Formato no soportado. Usa CSV, TXT, JSON o XLSX.",
    )


def upload_faq_entries(
    db: Session,
    *,
    company_id: UUID,
    filename: str,
    content: bytes,
    actor_user: User | None = None,
) -> AiFaqUploadResult:
    rows = _parse_faq_file(filename, content)
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo no contiene preguntas y respuestas validas.",
        )
    unique_map: dict[str, tuple[str, str]] = {}
    for question, answer in rows:
        key = question.strip().lower()
        if key:
            unique_map[key] = (question.strip(), answer.strip())
    cleaned_rows = list(unique_map.values())
    if len(cleaned_rows) > MAX_FAQ_ENTRIES_PER_TENANT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El archivo supera el maximo de {MAX_FAQ_ENTRIES_PER_TENANT} FAQs.",
        )

    existing = {
        entry.question.strip().lower(): entry
        for entry in list_faq_entries(db, company_id=company_id)
    }
    new_required = len([row for row in cleaned_rows if row[0].strip().lower() not in existing])
    if len(existing) + new_required > MAX_FAQ_ENTRIES_PER_TENANT:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Con las FAQs actuales, solo puedes agregar "
                f"{MAX_FAQ_ENTRIES_PER_TENANT - len(existing)} mas."
            ),
        )

    created = 0
    updated = 0
    for question, answer in cleaned_rows:
        key = question.strip().lower()
        current = existing.get(key)
        if current is None:
            current = AiFaqEntry(
                company_id=company_id,
                question=question,
                answer=answer,
                active=True,
            )
            db.add(current)
            existing[key] = current
            created += 1
            continue
        current.question = question
        current.answer = answer
        current.active = True
        updated += 1
    db.commit()
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.faq_entries.uploaded",
        entity_type="ai_faq_entry",
        entity_id=None,
        summary="FAQ file uploaded",
        metadata={"filename": filename, "created": created, "updated": updated, "total": len(cleaned_rows)},
    )
    return AiFaqUploadResult(total_read=len(cleaned_rows), created=created, updated=updated)


def list_interactive_templates(db: Session, *, company_id: UUID) -> list[AiInteractiveTemplate]:
    return list(
        db.scalars(
            select(AiInteractiveTemplate)
            .where(AiInteractiveTemplate.company_id == company_id)
            .order_by(AiInteractiveTemplate.updated_at.desc())
        )
    )


def create_interactive_template(
    db: Session,
    *,
    company_id: UUID,
    payload: AiInteractiveTemplateCreate,
    actor_user: User | None = None,
) -> AiInteractiveTemplate:
    action_key = payload.action_key.strip().lower()
    template = db.scalar(
        select(AiInteractiveTemplate).where(
            AiInteractiveTemplate.company_id == company_id,
            AiInteractiveTemplate.action_key == action_key,
        )
    )
    if template is None:
        template = AiInteractiveTemplate(company_id=company_id, action_key=action_key)
        db.add(template)
    template.name = payload.name.strip()
    template.template_type = payload.template_type
    template.body_text = payload.body_text.strip()
    template.footer_text = payload.footer_text
    template.button_text = payload.button_text
    template.section_title = payload.section_title
    template.options = [item.model_dump() for item in payload.options]
    template.usage_instruction = payload.usage_instruction.strip()
    template.trigger_mode = payload.trigger_mode
    template.trigger_fields = [
        field.strip().lower() for field in payload.trigger_fields if field.strip()
    ]
    template.active = payload.active
    db.commit()
    db.refresh(template)
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.interactive_template.saved",
        entity_type="ai_interactive_template",
        entity_id=template.id,
        summary="Interactive template saved",
        metadata={"action_key": template.action_key, "template_type": template.template_type},
    )
    return template


def get_interactive_template(
    db: Session, *, company_id: UUID, template_id: UUID
) -> AiInteractiveTemplate:
    template = db.scalar(
        select(AiInteractiveTemplate).where(
            AiInteractiveTemplate.company_id == company_id,
            AiInteractiveTemplate.id == template_id,
        )
    )
    if template is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Interactive template not found")
    return template


def update_interactive_template(
    db: Session,
    *,
    company_id: UUID,
    template_id: UUID,
    payload: AiInteractiveTemplateUpdate,
    actor_user: User | None = None,
) -> AiInteractiveTemplate:
    template = get_interactive_template(db, company_id=company_id, template_id=template_id)
    data = payload.model_dump(exclude_unset=True)
    if "action_key" in data and data["action_key"] is not None:
        data["action_key"] = data["action_key"].strip().lower()
        duplicate = db.scalar(
            select(AiInteractiveTemplate).where(
                AiInteractiveTemplate.company_id == company_id,
                AiInteractiveTemplate.action_key == data["action_key"],
                AiInteractiveTemplate.id != template_id,
            )
        )
        if duplicate is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="Ya existe una plantilla con ese action_key",
            )
    if "options" in data and data["options"] is not None:
        data["options"] = [item.model_dump() for item in payload.options or []]
    if "usage_instruction" in data and data["usage_instruction"] is not None:
        data["usage_instruction"] = data["usage_instruction"].strip()
    if "trigger_fields" in data and data["trigger_fields"] is not None:
        data["trigger_fields"] = [
            field.strip().lower() for field in data["trigger_fields"] if field.strip()
        ]
    for field, value in data.items():
        setattr(template, field, value)
    db.commit()
    db.refresh(template)
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.interactive_template.updated",
        entity_type="ai_interactive_template",
        entity_id=template.id,
        summary="Interactive template updated",
        metadata={"action_key": template.action_key, "template_type": template.template_type},
    )
    return template


def delete_interactive_template(
    db: Session,
    *,
    company_id: UUID,
    template_id: UUID,
    actor_user: User | None = None,
) -> None:
    template = get_interactive_template(db, company_id=company_id, template_id=template_id)
    template_id_value = template.id
    db.delete(template)
    db.commit()
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.interactive_template.deleted",
        entity_type="ai_interactive_template",
        entity_id=template_id_value,
        summary="Interactive template deleted",
        metadata={"action_key": template.action_key},
    )


def get_operational_config(db: Session, *, company_id: UUID, agent_id: UUID) -> AiOperationalConfigRead:
    agent = get_agent(db, company_id=company_id, agent_id=agent_id)
    operational_config = _agent_operational_config(db, company_id=company_id, agent=agent)
    return AiOperationalConfigRead(
        status=str(operational_config.get("status") or "draft"),
        version=int(operational_config.get("version") or 1),
        published_at=operational_config.get("published_at"),
        draft=operational_config.get("draft") if isinstance(operational_config.get("draft"), dict) else {},
        published=operational_config.get("published")
        if isinstance(operational_config.get("published"), dict)
        else operational_config.get("draft") if isinstance(operational_config.get("draft"), dict) else {},
    )


def update_operational_config(
    db: Session,
    *,
    company_id: UUID,
    agent_id: UUID,
    payload: dict,
    actor_user: User | None = None,
) -> AiAgent:
    agent = get_agent(db, company_id=company_id, agent_id=agent_id)
    fallback_timezone = _agent_timezone(db, company_id=company_id)
    rules = _agent_rules_copy(agent)
    rules["operational"] = build_operational_config(
        {"operational": payload},
        fallback_timezone=fallback_timezone,
    )
    validate_operational_config(rules["operational"])
    agent.rules = rules
    db.commit()
    db.refresh(agent)
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.operational_config.updated",
        entity_type="ai_agent",
        entity_id=agent.id,
        summary="Operational AI config updated",
        metadata={"status": agent.operational_status},
    )
    return agent


def publish_operational_config_for_agent(
    db: Session,
    *,
    company_id: UUID,
    agent_id: UUID,
    actor_user: User | None = None,
) -> AiAgent:
    agent = get_agent(db, company_id=company_id, agent_id=agent_id)
    fallback_timezone = _agent_timezone(db, company_id=company_id)
    rules = _agent_rules_copy(agent)
    operational_config = build_operational_config(rules, fallback_timezone=fallback_timezone)
    operational_config = publish_operational_config(operational_config)
    validate_operational_config(operational_config)
    rules["operational"] = operational_config
    agent.rules = rules
    db.commit()
    db.refresh(agent)
    _audit_ai_action(
        db,
        company_id=company_id,
        actor_user=actor_user,
        action="ai.operational_config.published",
        entity_type="ai_agent",
        entity_id=agent.id,
        summary="Operational AI config published",
        metadata={"status": agent.operational_status},
    )
    return agent


def simulate_operational_config(
    db: Session,
    *,
    company_id: UUID,
    agent_id: UUID,
    message: str,
    operational_config: dict | None = None,
) -> AiOperationalSimulationResponse:
    agent = get_agent(db, company_id=company_id, agent_id=agent_id)
    fallback_timezone = _agent_timezone(db, company_id=company_id)
    if operational_config is None:
        operational_config = _agent_operational_config(db, company_id=company_id, agent=agent)
    else:
        if not isinstance(operational_config, dict):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="La configuracion operativa debe ser un objeto valido.",
            )
        operational_config = build_operational_config(
            {"operational": operational_config},
            fallback_timezone=fallback_timezone,
        )
        validate_operational_config(operational_config)
    summary = simulation_summary(
        operational_config,
        timezone_name=fallback_timezone,
        message=message,
    )
    return AiOperationalSimulationResponse(**summary)
