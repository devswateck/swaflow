from __future__ import annotations

import argparse
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select

from app.companies.models import Company
from app.core.database import SessionLocal
from app.products.models import Product


@dataclass
class ProductRow:
    category: str
    name: str
    price: Decimal
    description: str
    image_files: list[str]


def normalize_key(value: str) -> str:
    ascii_value = (
        unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii").lower()
    )
    ascii_value = re.sub(r"_\d+$", "", ascii_value.strip())
    ascii_value = re.sub(r"[^a-z0-9]+", " ", ascii_value).strip()
    return re.sub(r"\s+", " ", ascii_value)


def parse_price(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    if isinstance(value, str):
        clean = value.replace("$", "").replace(".", "").replace(",", ".").strip()
        try:
            return Decimal(clean).quantize(Decimal("0.01"))
        except InvalidOperation as exc:
            raise ValueError(f"Precio invalido: {value!r}") from exc
    raise ValueError(f"Tipo de precio no soportado: {type(value)}")


def build_image_index(photos_dir: Path) -> dict[str, list[str]]:
    index: dict[str, list[str]] = {}
    for file_path in photos_dir.iterdir():
        if not file_path.is_file():
            continue
        if file_path.suffix.lower() not in {".png", ".jpg", ".jpeg", ".webp"}:
            continue
        key = normalize_key(file_path.stem)
        index.setdefault(key, []).append(file_path.name)
    for files in index.values():
        files.sort()
    return index


def read_rows(xlsx_path: Path, photos_dir: Path) -> list[ProductRow]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    raw_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    headers = [normalize_key(header) for header in raw_headers]

    wanted = {
        "categoria": None,
        "producto": None,
        "precio": None,
        "descripcion": None,
    }
    for index, header in enumerate(headers):
        if header in {"descripcion", "descripcion del producto"}:
            wanted["descripcion"] = index
        elif header in wanted:
            wanted[header] = index

    missing = [name for name, index in wanted.items() if index is None]
    if missing:
        raise ValueError(
            "No se encontraron columnas requeridas en el Excel: " + ", ".join(sorted(missing))
        )

    image_index = build_image_index(photos_dir)
    rows: list[ProductRow] = []

    for row_values in ws.iter_rows(min_row=2, values_only=True):
        name_value = row_values[wanted["producto"]]  # type: ignore[index]
        if not name_value:
            continue

        category_value = row_values[wanted["categoria"]]  # type: ignore[index]
        price_value = row_values[wanted["precio"]]  # type: ignore[index]
        description_value = row_values[wanted["descripcion"]]  # type: ignore[index]

        name = str(name_value).strip()
        category = str(category_value or "").strip()
        description = str(description_value or "").strip()
        price = parse_price(price_value)

        image_files = image_index.get(normalize_key(name), [])
        rows.append(
            ProductRow(
                category=category,
                name=name,
                price=price,
                description=description,
                image_files=image_files,
            )
        )

    return rows


def resolve_company_id(company_id: str | None) -> UUID:
    with SessionLocal() as db:
        if company_id:
            requested = UUID(company_id)
            company = db.scalar(select(Company).where(Company.id == requested))
            if company is None:
                raise ValueError(f"No existe company_id={company_id}")
            return requested

        companies = list(db.scalars(select(Company).where(Company.status == "active")))
        if len(companies) == 1:
            return companies[0].id
        if not companies:
            raise ValueError("No hay tenants activos en la base de datos.")

        printable = "\n".join(f"- {company.id} | {company.name}" for company in companies)
        raise ValueError(
            "Hay multiples tenants activos. Ejecuta con --company-id.\nDisponibles:\n" + printable
        )


def upsert_products(
    *,
    company_id: UUID,
    rows: list[ProductRow],
    currency: str,
    dry_run: bool,
) -> tuple[int, int]:
    created = 0
    updated = 0

    with SessionLocal() as db:
        for row in rows:
            existing = db.scalar(
                select(Product).where(Product.company_id == company_id, Product.name == row.name)
            )

            metadata = {
                "source": "excel_import",
                "category": row.category,
                "local_images": row.image_files,
                "primary_image": row.image_files[0] if row.image_files else None,
            }

            if existing is None:
                product = Product(
                    company_id=company_id,
                    name=row.name,
                    description=row.description,
                    sku=None,
                    price=row.price,
                    currency=currency,
                    status="active",
                    metadata_json=metadata,
                )
                db.add(product)
                created += 1
            else:
                existing.description = row.description
                existing.price = row.price
                existing.currency = currency
                existing.status = "active"
                current_metadata = (
                    existing.metadata_json if isinstance(existing.metadata_json, dict) else {}
                )
                current_metadata.update(metadata)
                existing.metadata_json = current_metadata
                updated += 1

        if dry_run:
            db.rollback()
        else:
            db.commit()

    return created, updated


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent.parent

    parser = argparse.ArgumentParser(
        description="Importa productos desde un Excel a la tabla products."
    )
    parser.add_argument("--company-id", help="UUID del tenant destino.")
    parser.add_argument(
        "--xlsx",
        default=str(repo_root / "productos" / "Productos.xlsx"),
        help="Ruta del archivo Excel.",
    )
    parser.add_argument(
        "--photos-dir",
        default=str(repo_root / "productos"),
        help="Carpeta con imagenes de productos.",
    )
    parser.add_argument("--currency", default="COP", help="Moneda a guardar en products.currency.")
    parser.add_argument("--dry-run", action="store_true", help="No persiste cambios en la BD.")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    photos_dir = Path(args.photos_dir).expanduser().resolve()

    if not xlsx_path.exists():
        raise SystemExit(f"No existe el archivo Excel: {xlsx_path}")
    if not photos_dir.exists():
        raise SystemExit(f"No existe la carpeta de fotos: {photos_dir}")

    company_id = resolve_company_id(args.company_id)
    rows = read_rows(xlsx_path, photos_dir)
    created, updated = upsert_products(
        company_id=company_id, rows=rows, currency=args.currency, dry_run=args.dry_run
    )

    mode = "DRY RUN" if args.dry_run else "IMPORTACION"
    print(
        f"{mode} OK | tenant={company_id} | leidos={len(rows)} | creados={created} | actualizados={updated}"
    )


if __name__ == "__main__":
    main()
